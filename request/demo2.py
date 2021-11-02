import requests
import json
import openpyxl as xl

'''
    批量测试IP专业接口并将测试结果保存到excel
'''

# URI = "136.192.124.177:30004"
URI = "135.64.134.201:30040"
headers = {
    'Content-Type': 'application/json;charset=UTF-8'
}
wb = xl.Workbook()
ws = wb.active

json_path = r'./西藏-考核-PON.json'


def my_request(method, url, data):
    # 此处用json=data  而不是data=data
    r = requests.request(method=method, url=url, json=data, headers=headers)
    # r.encoding = 'utf-8'
    # 自动获取编码
    r.encoding = r.apparent_encoding
    # text = r.text
    # result = json.loads(text)
    # result = json.loads(result)
    r_json = r.json()
    result = json.dumps(r_json, sort_keys=False, indent=4, separators=(',', ':'), ensure_ascii=False)
    # print(result)
    return result


def get_json():
    with open(json_path, 'r', encoding='utf-8') as file:
        return json.load(file)


def get_api_info():
    my_json = get_json()
    result_json = list()
    for item in my_json['item']:
        json_list = list()
        name = item['name']
        method = item['request']['method']
        url = str(item['request']['url']['raw']).replace("{{URI}}", URI)
        # url = f"http://{URI}/res/rpc/ctr/ip/switchbundleportquery"
        if str(method).lower() == 'post':
            url = url.split("?")[0] + "?"
            params = item['request']['body']['raw']
        else:
            # params = {'deviceId': '101.248.0.94'}
            params = url.split("?")[1]
        # data = {"extendInputParameter2": "[170.128.233.173]", "extendInputParameter1": "扩展入参1,预留扩展入参", "bundleId": ""}
        result = my_request(method, url, params)
        json_list.append(name)
        json_list.append(url)
        json_list.append(params)
        json_list.append(result)
        result_json.append(json_list)
    return result_json


def write_excel(row_num, item_list):
    column_num = 0
    result_code = 1000
    for item in item_list:
        column_num += 1
        result_code = item["resultCode"]
        ws.cell(row=row_num, column=column_num, value=item)
    ws.cell(row=row_num, column=column_num, value=result_code)


def exec_post():
    row_num = 0
    result_json = get_api_info()
    for item in result_json:
        row_num += 1
        write_excel(row_num, item)
    wb.save('result.xlsx')


exec_post()
