import openpyxl as xl

wb = xl.Workbook()
ws = wb.active

cellA = "A"
begin = cellA+"1"
mid = ":"
end = cellA+"50"
ws.merge_cells(begin+mid+end)
ws.merge_cells("B1:C5")
wb.save(r"C:\Users\420\Desktop\test.xlsx")

