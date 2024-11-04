import sys

import openpyxl as xl
from openpyxl.cell import MergedCell


def get_classes(ws) -> list:
    classes = []
    for row in ws.rows:
        rows = []
        for cell in row:
            if isinstance(cell, MergedCell):
                print(cell)
                break
            value = cell.value
            # print('value:',value)
            rows.append(value)
        classes.append(rows)
    return classes


def merge_list(col) -> list:
    lst = []
    for i in range(len(col)):
        value = col[i].value
        if value is None:
            continue
        lst.append(value)
        # print(value)
    # print(len(lst))
    return lst


def get_names(ws2) -> list:
    list_names = []
    bl = ws2['B']
    el = ws2['E']
    hl = ws2['H']
    list_names += merge_list(bl)
    list_names += merge_list(el)
    list_names += merge_list(hl)
    return list_names


def gen_result(ws3, title):
    for i in range(0, len(names)):
        count = 0
        name = ''
        # print('--name--',names[i])
        for j in range(1, len(classes)):
            for k in range(len(classes[j])):
                if str(names[i]).strip().replace(r'(', '').replace(r')', '') == str(classes[j][k]).strip():
                    if str(title[k]).__contains__('班主任'):
                        continue
                    if str(title[k]).__contains__('×'):
                        count += int(title[k][-1])
                    else:
                        count += 1
        # print(f'{names[i]},{count}')
        ws3.cell(i + 1, 1, names[i])
        ws3.cell(i + 1, 2, count)


def gen_spec_result() -> dict:
    for i in range(0, len(names)):
        name = str(names[i]).strip().replace(r'(', '').replace(r')', '')
        teacherDic: dict = result.get(name)
        dicClassCount = teacherDic.get(title)

        if result.get(name) is not None:
            pass
        else:
            result[name] = dict()

        for j in range(1, len(classes)):
            for k in range(len(classes[j])):
                if name == str(classes[j][k]).strip():
                    if dicClassCount is not None:
                        if str(title[k]).__contains__('班主任'):
                            continue
                        if str(title[k]).__contains__('×'):
                            dicClassCount += int(title[k][-1])
                        else:
                            dicClassCount += 1
                    else:
                        dicClassCount += 1
                    result[name][title] = dicClassCount
    return result


if __name__ == '__main__':

    # path = r'C:\Users\heave\Desktop\class.xlsx'
    path = r'C:\Users\heave\Desktop\8.26-1有统计结果.xlsx'
    path2 = r'C:\Users\heave\Desktop\任课教师名单.xlsx'

    # 课程表
    # path2 = input('请输入教师人员excel文件全路径：')
    # 任课教师表
    # path = input('请输入课程表excel文件全路径：')

    wb = xl.load_workbook(path)
    sheets = wb.sheetnames

    # ws = wb[sheets[sht_no]]

    wb2 = xl.load_workbook(path2)
    sheets2 = wb2.sheetnames
    ws2 = wb2[sheets2[0]]

    names = get_names(ws2)

    for sht_no in range(len(sheets)):
        sheet_title = sheets[sht_no] + '统计结果'
        ws = wb[sheets[sht_no]]

        wb.create_sheet(sheet_title, sht_no + 2)
        ws3 = wb[sheet_title]

        classes = get_classes(ws)
        title = classes[0]

        gen_result(ws3, title)
        result = gen_spec_result()

    # 保存结果 生成excel
    wb.save(path)
    # wb3.save('./lllllllllllllllllll.xlsx')
