import openpyxl as xl
from openpyxl import load_workbook

# 创建excel
create_wb = xl.Workbook()
# 获取默认的sheet
default_sheet = create_wb['Sheet']
# 创建新的sheet
new_ws = create_wb.create_sheet('new sheet')
# 删除默认的sheet
create_wb.remove(default_sheet)
# 构造数据
row = ['A11', 'A12', 'A13']
# 将数据添加到sheet
new_ws.append(row)
# 保存新的excel  执行此步才会最终创建一个excel
create_wb.save('new_excel.xlsx')

path = r'C:\Users\420\Desktop\kdgc\6月集团考核\RES Open API设计-IPRAN STN-20210520（皖通）-v1.0(1).xlsx'
# 加载excel文件
wb = xl.load_workbook(path)
# 激活excel
wb.active
sheets_name = wb.sheetnames
# 确定一个要操作的sheet
ws = wb[sheets_name[0]]
print(ws)


# 遍历行列 打印sheet
def for_sheet():
    for row in ws.rows:
        for cell in row:
            print(cell.value, end='\t')
        print()


cell_row1 = ''
cell_row2 = ''
cell_row3 = ''
cell_row4 = ''
cell_row5 = ''
max_row = ws.max_row
max_col = ws.max_column
for i in range(1, max_row):
    for j in range(1, 20):
        cell = ws.cell(i, j).value
        if j == 1 and cell is not None:
            cell_row1 = cell
        if j == 2 and cell is not None:
            cell_row2 = cell

        if cell is not None:
            print(cell, end='\t')
        else:
            if j == 1 and i > 1:
                print(cell_row1, end='\t')
            if j == 2:
                print(cell_row2, end='\t')
    print()
