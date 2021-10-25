import openpyxl as xl

'''
去重
'''

target_path = r'C:\Users\420\Desktop\kdgc\6月集团考核\附件1-集团考核API实指令模版-华为反馈V2版20210622.xlsx'
wb = xl.load_workbook(target_path)

sheet_names = wb.sheetnames

ws = wb[sheet_names[2]]
# print(sheet_names[2])
# 集团587个API
# 编码
target = ws['B']
# 名称
ta_name = ws['C']
# 待梳理的422个API
sample = ws['E']
count = 0
# target 字典
dic_ta = dict()
# sample 字典
dic_sa = dict()
# 587个API集合
lst = set()
lst3 = list()
# 422个API集合
lst2 = set()
for j in range(len(target)):
    value = target[j].value
    if value is not None:
        lst.add(target[j].value)
        lst3.append(target[j].value)
        # print(target[j].value)
print(len(lst))

for i in range(len(sample)):
    value = sample[i].value
    if value is not None:
        lst2.add(sample[i].value)
print(len(lst2))
cha = lst - lst2
# cha2 = lst2 - lst
# print(lst - lst2)
print(len(cha))
# print(len(cha2))

for i in cha:
    pass
    # print(i)
print('-----------------------------------------------------')

for t in range(len(lst3)):
    if lst3[t] in lst2:
        print(f'{t}==>{lst3[t]}')
    else:
        print(f'{t}==>{lst3[t]}==>diff')
