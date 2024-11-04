import openpyxl as xl
from openpyxl.cell import MergedCell

'''
    根据任课老师和课表 查询课程冲突的老师名单
'''


# 周次及课程顺序
def read_side_value():
    week_order = [i for i in range(1, 6)] * 10
    course_order = [i for i in range(1, 8)]
    return week_order, course_order


def read_merged_cell_value(ws, max_row, start_row, col):
    # 自下而上 逐行遍历 最多遍历到课程标题行前一行
    for i in range(max_row, start_row, -1):
        # 获取当前单元格值
        cell = ws.cell(i, col)._value
        # 如果当前单元格值不存在 向上取值
        if cell is not None:
            return cell


# 读教师表
def read_teacher_course_map():
    all_grade_list = []
    # 1-8行为1年级
    grade_one_list = read_grade_x(1, 8, 1)
    # 9-15行2年级
    grade_two_list = read_grade_x(9, 15, 2)
    # 16-23行3年级
    grade_three_list = read_grade_x(16, 23, 3)
    # 24-30行4年级
    grade_four_list = read_grade_x(24, 30, 4)
    # 31-37行5年级
    grade_five_list = read_grade_x(31, 37, 5)
    # 38-44行6年级
    grade_six_list = read_grade_x(38, 44, 6)
    all_grade_list += grade_one_list
    all_grade_list += grade_two_list
    all_grade_list += grade_three_list
    all_grade_list += grade_four_list
    all_grade_list += grade_five_list
    all_grade_list += grade_six_list
    # print(all_grade_list)
    return all_grade_list


# 获取每个年级教师任课信息
def read_grade_x(start_row, end_row, grade_no) -> list:
    grade_map_list = []
    last_sheet = wb2.sheetnames[-1]
    ws = wb2[last_sheet]
    max_col = ws.max_column
    # +1为了去除表头的课程名称行和侧边的班级列
    for row in range(start_row + 1, end_row + 1):
        for col in range(2, max_col + 1):
            # 老师名称
            cell = ws.cell(row, col)._value
            # 课程名称
            course_name = str(ws.cell(start_row, col).value).replace('\n', '').strip()
            # 如果课程名称为空 说明是空列 该列无课程
            if course_name == 'None':
                continue
            # 如果老师为空说明是合并单元格 取第一个非空的值
            if cell is None:
                cell = read_merged_cell_value(ws, row, start_row, col)
            # 一个单元格可能有多个老师
            teachers = str(cell).split('/')
            for teacher in teachers:
                it = dict()
                it['name'] = teacher
                it['class'] = str(grade_no) + '.' + str(row - start_row)
                it['course'] = course_name
                grade_map_list.append(it)
    return grade_map_list


# 读课程表
def read_course_map() -> list:
    course_list = []
    sheets: list = wb.sheetnames
    # print(f'sheets:{sheets}')

    wko, coo = read_side_value()
    # print(wko)
    # print(coo)

    for i in range(len(sheets)):
        wsi = wb[sheets[i]]
        rows = wsi.max_row
        cols = wsi.max_column
        # print(f'行数：{rows},列数：{cols}')
        for r in range(4, rows + 1):
            # 第8行为分割行
            if r == 8:
                continue
            for c in range(3, cols + 1):
                # ._value可读取合并单元格和正常单元格 .value只能读取正常单元格
                cell = wsi.cell(r, c)._value
                it = dict()
                # 课程名称 去除换行符
                it['course'] = str(cell).replace('\n', '')
                # 年级
                # it['grade'] = wko[c - 3]
                # 第几节课
                if r > 8:
                    # 第8行为分割行 不计课程顺序 故行数加1 但课课程顺序不能加1
                    it['order'] = coo[r - 5]
                else:
                    it['order'] = coo[r - 4]
                # 班级和年级 第一个sheet(i值)为一年级 以此类推 每隔五列为一个班级
                it['class'] = str(i + 1) + '.' + str((c - 3) // 5 + 1)
                # 星期几的课程
                it['weekDay'] = wko[c - 3]
                course_list.append(it)
                # print(f'{r}行{c}列：{cell}', end='\t')
            # print('\n')
        # print('-----------sheet--------------')
    return course_list


# 获取所有老师及其教的课程
def get_all_teacher_course_info() -> list:
    all_teacher_list = []

    for teacher in teacher_course_list:
        teacher_name: str = teacher.get('name')
        teacher_class: str = teacher.get('class')
        teacher_course: str = teacher.get('course')
        for course in course_list:
            course_name: str = course.get('course')
            course_class: str = course.get('class')
            course_weekDay: str = course.get('weekDay')
            course_order: str = course.get('order')
            # 两表班级相同 课程名相同 则认为该老师教对应的这节课
            if teacher_class == course_class and teacher_course.__contains__(course_name):
                teacher_course_info = dict()
                teacher_course_info.update(course)
                # teacher_course_info['count'] = 1
                teacher_course_info['name'] = teacher_name

                all_teacher_list.append(teacher_course_info)
                # print(f'{teacher_course_info}')
    return all_teacher_list


def compare_teacher_course():
    unique_info_map = dict()
    for mp in all_teacher_list:
        name = mp.get('name')
        week = mp.get('weekDay')
        order = mp.get('order')
        course = mp.get('course')
        # mp['duplicate'] = []
        unique_key = name + '-' + str(week) + '-' + str(order)
        if unique_info_map.keys().__contains__(unique_key) and unique_info_map[unique_key]['course'] != course:
            # print(f'course:{unique_info_map[unique_key]["course"]} == {course}')
            # duplicate_course = mp['duplicate']
            # unique_info_map[unique_key]['duplicate'].append(str(mp))
            unique_info_map[unique_key]['course'] += ', ' + mp['course']
            unique_info_map[unique_key]['class'] += ', ' + mp['class']
            # mp['duplicate'] = duplicate_course
            print(f'{unique_key}==课程冲突=={unique_info_map[unique_key]}')
            continue
        unique_info_map[unique_key] = mp

        # if name == teacher_name and week == course_weekDay and course_order == order:
        #     teacher_course_info['count'] += 1
        #     teacher_course_info['course'] += ',' + course
        #     duplicate_teacher_list.append(teacher_course_info)


if __name__ == '__main__':
    teacher_course_path = r'./files/朝阳小学任课表2024年上.xlsx'
    course_path = r'./files/副本2023-2024学年度 日照市朝阳小学总课程表.xlsx'

    wb = xl.load_workbook(course_path)
    wb2 = xl.load_workbook(teacher_course_path)

    # 课程表列表
    course_list = read_course_map()
    # 任课教师列表
    teacher_course_list = read_teacher_course_map()

    # print(course_list)
    # print(teacher_course_list)
    all_teacher_list = get_all_teacher_course_info()

    compare_teacher_course()
    # print(duplicate_teacher)
