import openpyxl as xl
import json


def read_json():
    path = r"C:\Users\heave\Desktop\DeskTop\orgList.json"

    with open(path, 'r', encoding='utf-8') as fs:
        jstr = json.load(fs)
        return jstr


def convert_json():
    result_path = r'./result.xlsx'
    wb = xl.Workbook()
    ws = wb.active

    jstr = read_json()
    row_num = 0

    for node in jstr.get('hrorgs'):
        row_num += 1
        # print(node)
        # node = json.load(node)
        org_name = node.get('orgName')
        org_code = node.get('orgCode')
        pk_org = node.get('pk_org')
        pk_fatherorg = node.get('pk_fatherorg')
        ws.cell(column=1, row=row_num, value=org_code)
        ws.cell(column=2, row=row_num, value=pk_org)
        ws.cell(column=3, row=row_num, value=pk_fatherorg)
        ws.cell(column=4, row=row_num, value=org_name)
        ws.cell(column=5, row=row_num, value=row_num)
    wb.save(result_path)


if __name__ == '__main__':
    # jstr = read_json()
    # jArray = jstr.get('hrorgs')
    # print(jArray)
    convert_json()
