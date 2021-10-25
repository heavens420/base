import openpyxl as xl

'''
    单元格合并
'''

path = r'../docx/总表.xlsx'
# path = r"C:\Users\420\Desktop\test.xlsx"
wb = xl.load_workbook(path)
sheets = wb.sheetnames
ws = wb[sheets[0]]

columnA = ws["A"]
begin_num = 1
end_num = 1
mid = ":"
merge_num = 0


def merge_row(column):
    global begin_num
    global end_num
    global merge_num
    A = "A"
    length = len(column)
    # 只有两行或以上相同内容的连续单元格按行合并，总共就一行就别来凑热闹了
    if length > 1:
        # 遍历所有行 从第二行开始遍历：因为遍历第一行毫无意义 我们要合并两行或以上内容相同行
        for i in range(1, length):
            # print(column[i].value, end='\n')
            # 总行数加一 行数后移
            end_num += 1
            # 如果当前行和前一行单元格内容一样 那就合并了
            if column[i].value != column[i - 1].value:
                # 又多一个合并行
                merge_num += 1
                # 执行合并行
                ws.merge_cells(A + str(begin_num) + mid + A + str(end_num - 1))
                # 下一个合并行的开始位置 为上一个合并行结束位置再加1 （由于上一合并行末行为 end_num - 1 所以这里就不用再加1了）
                begin_num = end_num
        # 最后的多行合并：最后一行的后一行（即为length+1行 ）为空 此时最后一行后的一行无法被遍历
        # 即无法触发if条件 故最后一个待合并的多行会出问题  需单独处理
        ws.merge_cells(A + str(begin_num) + mid + A + str(end_num))
    # 保存合并修改
    wb.save(path)


merge_row(columnA)
print(f'begin_num = {begin_num}\nend_num = {end_num}\t\n')
print(f'总共有{merge_num}个合并行')
