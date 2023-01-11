from docx import Document
import re
import openpyxl as xl

wb = xl.Workbook()
ws = wb.active

# path = r'C:\Users\420\Desktop\kdgc\西藏阿里PON专业OLT性能指标采集配置过程.docx'
path = r'C:\Users\420\Desktop\xxx.docx'
obj = Document(path)
row_num = 1

for item in obj.paragraphs:
    # if re.match("^Heading \d+$", content.style.name):
    if item is not None or item != '':
        column_num = 0
        if re.match("^Heading 3", item.style.name):
            column_num += 1
            ws.cell(row=row_num, column=column_num, value=item.text)
        item = str(item.text)
        if str(item).startswith("接口名称"):
            column_num += 1
            # print(f'{item}')
            ws.cell(row=row_num, column=column_num, value=item)
        if str(item).startswith("Path"):
            column_num += 1
            # print(f'{item}')
            ws.cell(row=row_num, column=column_num, value=item)
        if str(item).startswith("Method"):
            column_num += 1
            # print(f'{item}')
            ws.cell(row=row_num, column=column_num, value=item)
            row_num += 1
wb.save('result1.xlsx')
