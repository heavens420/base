from docx import Document
import openpyxl as xl

wb = xl.Workbook()
ws = wb.active

# path = r'C:\Users\420\Desktop\中国电信新一代云网运营业务系统采控中心RES open API接口规范-传输专业分册-0826.mydocx'
path = r'C:\Users\420\Desktop\中国电信新一代云网运营业务系统采控中心RES open API接口规范-传输专业分册-0826 - 副本.mydocx'
obj = Document(path)

result_list = list()
count = 1
for table in obj.tables:
    item = table.cell(0, 0).text
    length = len(table.rows)
    column_length = len(table.columns)
    if "参数名称".__eq__(str(item).strip()) or "名称".__eq__(str(item).strip()) and column_length > 4:
        title = list()
        title_flag = count % 3
        if title_flag == 1:
            title.append("Headers")
        if title_flag == 2:
            title.append("Body")
        if title_flag == 0:
            title.append("返回数据")
        count += 1
        # result_list.append(title)

        for row in range(length):
            table_list = list()
            for column in range(column_length):
                item = table.cell(row, column).text
                # print(item, end='\t')
                table_list.append(item)
            # print('', end='\n')
            result_list.append(table_list)
            # print(table_list)

for row in range(len(result_list)):
    for column in range(len(result_list[row])):
        ws.cell(row=row+1, column=column+1, value=result_list[row][column])
wb.save('result.xlsx')
